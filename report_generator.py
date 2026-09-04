"""
Report Generator for AI Risk Assessment Tool
Generates professional Excel reports with executive summaries
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List


class ReportGenerator:
    """Generates professional Excel reports for AI risk assessments"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def generate_report(self, assessment: Dict, output_path: str) -> None:
        """
        Generate comprehensive Excel report from assessment results
        
        Args:
            assessment: Assessment results dictionary
            output_path: Path to save the Excel file
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "AI Risk Assessment"
        
        # Add executive summary
        self._add_executive_summary(assessment)
        
        # Add system classification
        self._add_system_classification(assessment)
        
        # Add compliance status
        self._add_compliance_status(assessment)
        
        # Add risk assessment
        self._add_risk_assessment(assessment)
        
        # Add control gaps
        self._add_control_gaps(assessment)
        
        # Add recommendations
        self._add_recommendations(assessment)
        
        # Save the workbook
        self.workbook.save(output_path)
    
    def _add_executive_summary(self, assessment: Dict) -> None:
        """Add executive summary section"""
        row = 1
        
        # Title
        self._add_title(row, "AI Risk Assessment Executive Summary")
        row += 2
        
        # System Information
        self._add_section_header(row, "System Information")
        row += 1
        self._add_text(row, 1, f"System Name: {assessment['system_name']}")
        row += 1
        self._add_text(row, 1, f"Assessment Date: {assessment['assessment_date']}")
        row += 1
        self._add_text(row, 1, f"System Type: {assessment['classification']['system_name']}")
        row += 1
        self._add_text(row, 1, f"Risk Factor: {assessment['composite_risk_factor']}")
        row += 2
        
        # Overall Risk Score
        self._add_section_header(row, "Overall Risk Assessment")
        row += 1
        self._add_text(row, 1, f"Risk Score: {assessment['risk_score']}")
        self._add_text(row, 2, f"Severity: {assessment['severity']}")
        self._add_text(row, 3, f"Status: {assessment['risk_description']}")
        row += 2
        
        # Compliance Overview
        self._add_section_header(row, "Compliance Overview")
        row += 1
        self._add_text(row, 1, f"Control Coverage: {assessment['control_coverage']}%")
        self._add_text(row, 2, f"Required Controls: {len(assessment['required_controls'])}")
        self._add_text(row, 3, f"Implemented: {len(assessment['implemented_controls'])}")
        self._add_text(row, 4, f"Gaps: {len(assessment['compliance_gaps'])}")
        row += 2
    
    def _add_system_classification(self, assessment: Dict) -> None:
        """Add system classification details"""
        row = self.worksheet.max_row + 2
        
        self._add_section_header(row, "System Classification Details")
        row += 1
        
        classification = assessment['classification']
        self._add_text(row, 1, f"Type: {classification['system_name']}")
        row += 1
        self._add_text(row, 1, f"Risk Factor: {classification['risk_factor']}")
        row += 1
        self._add_text(row, 1, f"Control Categories: {', '.join(classification['control_categories'])}")
        row += 1
        self._add_text(row, 1, f"Confidence: {classification['classification_confidence']:.0%}")
        row += 1
        self._add_text(row, 1, f"Description: {classification['description']}")
        row += 2
    
    def _add_compliance_status(self, assessment: Dict) -> None:
        """Add compliance status breakdown"""
        row = self.worksheet.max_row + 2
        
        self._add_section_header(row, "Compliance Status Breakdown")
        row += 1
        
        # Headers
        self._add_header(row, 1, "Metric")
        self._add_header(row, 2, "Value")
        self._add_header(row, 3, "Status")
        row += 1
        
        # Data
        self._add_text(row, 1, "Control Coverage")
        self._add_text(row, 2, f"{assessment['control_coverage']}%")
        self._add_status(row, 3, assessment['control_coverage'] >= 80)
        row += 1
        
        self._add_text(row, 1, "Required Controls")
        self._add_text(row, 2, str(len(assessment['required_controls'])))
        self._add_text(row, 3, "Total")
        row += 1
        
        self._add_text(row, 1, "Implemented Controls")
        self._add_text(row, 2, str(len(assessment['implemented_controls'])))
        self._add_text(row, 3, "Completed")
        row += 1
        
        self._add_text(row, 1, "Compliance Gaps")
        self._add_text(row, 2, str(len(assessment['compliance_gaps'])))
        self._add_status(row, 3, len(assessment['compliance_gaps']) == 0)
        row += 2
    
    def _add_risk_assessment(self, assessment: Dict) -> None:
        """Add detailed risk assessment"""
        row = self.worksheet.max_row + 2
        
        self._add_section_header(row, "Risk Assessment Details")
        row += 1
        
        # Risk Score
        self._add_text(row, 1, f"Overall Risk Score: {assessment['risk_score']}")
        self._add_text(row, 2, f"Severity Level: {assessment['severity']}")
        self._add_text(row, 3, f"Description: {assessment['risk_description']}")
        row += 2
        
        # Risk Factors
        self._add_section_header(row, "Key Risk Factors")
        row += 1
        for factor in assessment['risk_factors']:
            self._add_text(row, 1, f"• {factor}")
            row += 1
        row += 1
        
        # Additional Risk Factors
        self._add_section_header(row, "Additional Risk Factors")
        row += 1
        for factor, value in assessment['additional_risk_factors'].items():
            self._add_text(row, 1, f"{factor.replace('_', ' ').title()}: {value}")
            row += 1
        row += 2
    
    def _add_control_gaps(self, assessment: Dict) -> None:
        """Add control gaps breakdown"""
        if not assessment['compliance_gaps']:
            return
            
        row = self.worksheet.max_row + 2
        
        self._add_section_header(row, "Compliance Gaps Analysis")
        row += 1
        
        # Headers
        self._add_header(row, 1, "Control ID")
        self._add_header(row, 2, "Gap Description")
        self._add_header(row, 3, "Severity")
        row += 1
        
        # Data
        for gap in assessment['compliance_gaps']:
            parts = gap.split(": ", 1)
            if len(parts) == 2:
                self._add_text(row, 1, parts[0])
                self._add_text(row, 2, parts[1])
                self._add_text(row, 3, "HIGH")
            else:
                self._add_text(row, 1, gap)
                self._add_text(row, 2, "Analysis required")
                self._add_text(row, 3, "MEDIUM")
            row += 1
        row += 2
    
    def _add_recommendations(self, assessment: Dict) -> None:
        """Add recommendations section"""
        row = self.worksheet.max_row + 2
        
        self._add_section_header(row, "Recommendations")
        row += 1
        
        # Headers
        self._add_header(row, 1, "Priority")
        self._add_header(row, 2, "Recommendation")
        row += 1
        
        # Data
        for i, recommendation in enumerate(assessment['recommendations'], 1):
            priority = "HIGH" if i <= 3 else "MEDIUM"
            self._add_text(row, 1, priority)
            self._add_text(row, 2, recommendation)
            row += 1
    
    def _add_title(self, row: int, text: str) -> None:
        """Add formatted title"""
        cell = self.worksheet.cell(row=row, column=1, value=text)
        cell.font = Font(size=16, bold=True, color="1E3A5F")
        cell.alignment = Alignment(horizontal="center")
        self.worksheet.merge_cells(f"A{row}:C{row}")
    
    def _add_section_header(self, row: int, text: str) -> None:
        """Add section header"""
        cell = self.worksheet.cell(row=row, column=1, value=text)
        cell.font = Font(size=12, bold=True, color="1E3A5F")
        cell.fill = PatternFill(start_color="E8E4D9", end_color="E8E4D9", fill_type="solid")
        self.worksheet.merge_cells(f"A{row}:C{row}")
    
    def _add_header(self, row: int, col: int, text: str) -> None:
        """Add table header"""
        cell = self.worksheet.cell(row=row, column=col, value=text)
        cell.font = Font(size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    def _add_text(self, row: int, col: int, text: str) -> None:
        """Add text cell"""
        cell = self.worksheet.cell(row=row, column=col, value=text)
        cell.font = Font(size=10)
        cell.alignment = Alignment(wrap_text=True)
    
    def _add_status(self, row: int, col: int, is_good: bool) -> None:
        """Add status indicator"""
        cell = self.worksheet.cell(row=row, column=col, value="✓" if is_good else "✗")
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(
            start_color="10B981" if is_good else "F59E0B",
            end_color="10B981" if is_good else "F59E0B",
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")
    
    def _format_columns(self) -> None:
        """Auto-format column widths"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width


# Example usage
if __name__ == "__main__":
    generator = ReportGenerator()
    
    # Sample assessment data
    sample_assessment = {
        "system_name": "Test AI System",
        "assessment_date": "2026-09-04T04:45:00",
        "classification": {
            "system_name": "Generative AI",
            "risk_factor": 2.0,
            "control_categories": ["all"],
            "classification_confidence": 0.9,
            "description": "AI system that generates new content"
        },
        "composite_risk_factor": 2.0,
        "required_controls": ["AI_GOV_001", "AI_GOV_002"],
        "implemented_controls": ["AI_GOV_001"],
        "control_coverage": 50.0,
        "compliance_gaps": ["AI_GOV_002: AI roles and responsibilities defined"],
        "risk_score": 75.0,
        "severity": "HIGH",
        "risk_description": "Action required within 30 days",
        "risk_factors": ["Low control coverage", "High system risk factor"],
        "additional_risk_factors": {
            "data_sensitivity": 1.0,
            "user_impact": 1.0,
            "regulatory_requirements": 1.0,
            "autonomy_level": 1.0
        },
        "recommendations": [
            "Develop comprehensive remediation plan",
            "Implement additional monitoring and controls",
            "Establish timeline for compliance"
        ]
    }
    
    generator.generate_report(sample_assessment, "ai_risk_assessment_report.xlsx")
    print("Report generated: ai_risk_assessment_report.xlsx")